import os
import openpyxl
import requests
from bs4 import BeautifulSoup

xlsx = 'vprognoze.xlsx'
url_txt = 'url.txt'
bank_txt = 'bank.txt'
all_win = []
all_lost = []


def bank_fun():
    my_file = open(bank_txt, "w+")
    my_file.write('1000')
    my_file.close()


def url_fun():
    my_file = open(url_txt, "w+")
    my_file.write('https://vprognoze.ru/?cid=202202&action=rating&do=shop&uid=1663824')
    my_file.close()


def xlsx_fun():
    filepath = xlsx
    wb = openpyxl.Workbook()
    wb.save(filepath)


def clear(bank):
    filepath = xlsx
    wb = openpyxl.Workbook()
    wb.save(filepath)
    wb = openpyxl.load_workbook(xlsx)
    sheet = wb.active
    data = (
        ()
    )
    for row in data:
        sheet.append(row)
    wb.save(xlsx)

    wb = openpyxl.load_workbook(xlsx)
    sheet = wb.active
    sheet.append(('Дата', 'Время', 'Команды', 'Лиги', 'Счет', 'Тотал', 'Статус', 'Коэффициент', 'Процент от банка, в %',
                  'Процент от банка, в руб', "Выйгрыш", "Проигрыш", "Итог", "БАНК: " + str(bank)))
    wb.save(xlsx)


def append(*args):
    wb = openpyxl.load_workbook(xlsx)
    sheet = wb.active
    for row in args:
        sheet.append(row)
    wb.save(xlsx)



def userid(url, banks):
    bank = banks
    url = url.replace('.ru/', '.kz/')
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    for ss in soup.find_all('div', class_='section__body'):
        for j in ss.find_all('div', class_='mini-tip-list'):
            for i in reversed(j.find_all('div', class_='mini-tip')):
                data = i.find('div', class_='ui-date__day').text
                time = i.find('div', class_='ui-date__hour').text
                try:
                    commands_teams = i.find('div', class_='mini-tip__teams').text
                except:
                    try:
                        commands_teams = i.find('div', class_='mini-tip__sport').text
                    except:
                        commands_teams = i.find('div', class_='mini-tip__express').text

                try:
                    commands_leages = i.find('div', class_='mini-tip__league').text
                except:
                    commands_leages = ''

                stake = i.find('div', class_='mini-tip__stake').text.replace("Сумма", "").replace("+", "").replace("-",
                                                                                                                   "").replace(
                    "%", "")  # Процент от банка
                status = i.find('div', class_='mini-tip__stake').text.replace("Сумма", "")[:1]

                try:
                    i.find('div', class_='mini-tip__bet').find("span").find('sup').extract()
                except:
                    pass

                try:
                    total_kof = i.find('div', class_='mini-tip__bet').find("span").text.replace("@", "").replace(" ",
                                                                                                                 "")  # коэффициент
                except:
                    total_kof = i.find('div', class_='mini-tip__bet').text.replace("@", "").replace(" ",
                                                                                                    "")  # коэффициент

                try:
                    total = i.find('div', class_='mini-tip__bet').find('div',
                                                                       class_='mini-tip__bet-inner').text.replace('@',
                                                                                                                  '').replace(
                        str(total_kof), '')  # тотал
                except:
                    total = ' '

                try:
                    score = i.find('div', class_='mini-tip__score').text  # счет
                except:
                    score = ' '
                try:
                    # действия
                    stake_rub = float(bank) * float(stake) / 100
                    if status == "+":
                        marja_win = int(float(stake_rub) * float(total_kof))
                        marja_lost = ' '
                        bank = int(float(bank) + float(marja_win))
                    else:
                        marja_win = ' '
                        marja_lost = stake_rub
                        bank = int(float(bank) - float(marja_lost))
                except:
                    marja_win = ' '
                    marja_lost = ' '
                    stake_rub = ' '

                    # запись
                stake = stake.replace(".", ",")
                total_kof = total_kof.replace(".", ",")
                bank = int(bank)
                print(data, time, commands_teams, commands_leages, score, total, status, total_kof, stake, stake_rub,
                      marja_win, marja_lost, bank)
                append([data, time, commands_teams, commands_leages, score, total, status, total_kof, stake, stake_rub,
                        marja_win, marja_lost, bank])

                print('\n')


def main():
    if not os.path.exists(bank_txt):
        bank_fun()
    if not os.path.exists(url_txt):
        url_fun()
    if not os.path.exists(xlsx):
        xlsx_fun()

    with open(bank_txt) as f:
        for i in f:
            banks = int(str(i.replace("\n", "")).replace(" ", ""))
    print(banks)
    clear(banks)
    with open(url_txt) as f:
        for i in f:
            url = str(i.replace("\n", "")).replace(".ru", ".kz")
            print(url)
            userid(url, banks)
            append((" ", " "))


if __name__ == "__main__":
    main()
