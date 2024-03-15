import configparser
import asyncio
import json
import time

import openpyxl
import requests
from fake_useragent import UserAgent
import aiohttp
from bs4 import BeautifulSoup


def clear():
    filepath = "15.xlsx"
    wb = openpyxl.Workbook()
    wb.save(filepath)
    wb = openpyxl.load_workbook("15.xlsx")
    sheet = wb.active
    data = (
        ()
    )
    for row in data:
        sheet.append(row)
    wb.save('15.xlsx')



def append(*args):
    wb = openpyxl.load_workbook("5.xlsx")
    sheet = wb.active
    for row in args:
        sheet.append(row)
    wb.save('5.xlsx')
s2 = ['Ириска', 'Какао', 'Кактус', 'Карамель', 'Кардамон', 'Квас', 'Киви', 'Кислота', 'Кислый', 'Клевер', 'Кленовый', 'Клубника', 'Клюква', 'Кокос', 'Кола', 'Конфетный', 'Корица', 'Кофе', 'Крыжовник', 'Кукуруза', 'Лаванда', 'Лайм', 'Лакрица', 'Леденцы', 'Лемонграсс', 'Лимон', 'Лимонад', 'Личи', 'Малина', 'Манго', 'Мандарин', 'Маракуйя', 'Мармелад', 'Марула', 'Мастика', 'Мед', 'Мелисса', 'Ментол', 'Микс', 'Миндаль', 'Молоко', 'Морковь', 'Мороженое', 'Морошка', 'Мультифрукт', 'Мясо', 'Мята', 'Нектарин', 'Новогодний', 'Нуга', 'Облепиха', 'Овсяные хлопья', 'Огурец', 'Орех', 'Орчата', 'Папайя', 'Перец', 'Персик', 'Пиво', 'Питахайя', 'Помело', 'Попкорн', 'Портвейн', 'Пряный', 'Ревень', 'Роза', 'Ром', 'Рутбир', 'Сакура', 'Сандал', 'Семечки', 'Сидр', 'Слива', 'Сливочный', 'Смородина', 'Специи', 'Сыр', 'Табачный', 'Тархун', 'Текила', 'Тирамису', 'Томат', 'Травяной', 'Тыква', 'Фейхоа', 'Фисташки', 'Фруктовый', 'Хвойный', 'Хлеб', 'Холодок', 'Цветочный', 'Цитрусовый', 'Чай', 'Черника', 'Чернослив', 'Чеснок', 'Чизкейк', 'Шалфей', 'Шампанское', 'Шафран', 'Шоколад', 'Эвкалипт', 'Энергетик', 'Яблоко', 'Ягодный']
s1 = ['95', '76', '50', '139', '101', '78', '45', '129', '119', '87', '105', '146', '81', '126', '102', '54', '131', '136', '108', '116', '128', '1', '19', '71', '15', '106', '93', '103', '40', '121', '79', '51', '14', '62', '100', '6', '60', '156', '46', '47', '8', '30', '23', '94', '155', '57', '125', '123', '149', '18', '91', '113', '34', '24']

# парсинг категории(url) товаров(ссылок)
def pars_tovar():

    async def get_page_data(session, logo, name, line, brand, urls):
        ua = UserAgent()
        header = {'User-Agent': str(ua.random)}

        async with session.get(urls, ssl=False, headers=header) as resp:

                resp_text = await resp.text()
                soup = BeautifulSoup(resp_text, 'lxml')
                try:
                    text_brand = str(soup.find('div', class_='object_brand').text).replace('\n', ' ')
                except:
                    text_brand = 'Нету'
                try:
                    text_line = str(soup.find('div', class_='object_line').text).replace('\n', ' ')
                except:
                    text_line = 'Нету'

                append((logo, name, line, brand, urls, text_brand, text_line))


    async def load_site_data():
        async with aiohttp.ClientSession(trust_env=True) as session:
            tasks = []
            for count, item in enumerate(s1):
                print(f'{count}-{len(s1)}')
                append((s2[count], '', '', '', '', '', '', '', '', '', ''))
                append(('Лого-сылка', 'Название', 'Линейка', 'Бренд', 'Ссылка на товар', 'Описание бренда',
                        'Описание Линейки'))

                for i in range(0, 300, 20):
                    url = 'https://htreviews.org/getData?r=flavor&s=rating&d=desc&t=' + str(item) + '&o=' + str(
                        i) + '&action=tobaccos'

                    try:
                        response = requests.get(url)
                        data = response.json()
                        for i in data:
                            logo = str('https://htreviews.org/' + str(i['media']))  # файл
                            name = str(i['name'])  # имя
                            line = str(i['line'])  # line
                            brand = str(i['brand'])  # brand
                            urls = str('https://htreviews.org/tobaccos/' + str(i['slug']))  # ссылка

                            task = asyncio.create_task(get_page_data(session, logo, name, line, brand, urls))
                            tasks.append(task)
                    except:
                        pass
                    await asyncio.gather(*tasks)


    asyncio.run(load_site_data())



if __name__ == "__main__":
    pars_tovar()
